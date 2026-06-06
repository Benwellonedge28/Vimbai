"""
PDF and Excel Export Service for FinAcc Reporting
Provides comprehensive export functionality for financial reports
"""

import io
import json
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
from decimal import Decimal
import asyncio

# For PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        Image, PageBreak, TableOfContents, PageTemplate, BaseDocTemplate
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
    from reportlab.pdfgen import canvas
    from reportlab.platypus.flowables import HRFlowable
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# For Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# For additional formatting
try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


class ExportService:
    """Service for exporting financial reports to PDF and Excel formats"""

    def __init__(self, company_name: str = "FinAcc", logo_path: Optional[str] = None):
        self.company_name = company_name
        self.logo_path = logo_path

        # Check availability of required libraries
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    async def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        title: str,
        report_type: str = "Financial Report",
        columns: Optional[List[str]] = None,
        include_summary: bool = True,
        page_size: str = "A4",
        orientation: str = "portrait"
    ) -> bytes:
        """
        Export data to a PDF document

        Args:
            data: List of dictionaries containing report data
            title: Report title
            report_type: Type of report (e.g., "Income Statement", "Balance Sheet")
            columns: Optional list of column names to include (defaults to all keys)
            include_summary: Whether to include summary statistics
            page_size: Page size (A4, letter)
            orientation: Page orientation (portrait, landscape)

        Returns:
            PDF document as bytes
        """
        buffer = io.BytesIO()

        # Determine page size
        if page_size.lower() == "letter":
            page_format = letter
        else:
            page_format = A4

        # Determine orientation
        if orientation.lower() == "landscape":
            page_format = landscape(page_format)

        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_format,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )

        # Build document content
        elements = []

        # Header
        elements.extend(self._create_pdf_header(title, report_type))

        # Table data
        if columns is None and data:
            columns = list(data[0].keys())

        if data and columns:
            elements.extend(self._create_pdf_table(data, columns))

        # Summary section
        if include_summary and data:
            elements.extend(self._create_pdf_summary(data))

        # Footer with page numbers
        doc.build(elements, onFirstPage=self._add_page_footer, onLaterPages=self._add_page_footer)

        buffer.seek(0)
        return buffer.getvalue()

    def _create_pdf_header(self, title: str, report_type: str) -> List:
        """Create PDF header section"""
        elements = []
        styles = getSampleStyleSheet()

        # Company name
        company_style = ParagraphStyle(
            'CompanyName',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a365d'),
            alignment=TA_CENTER,
            spaceAfter=6
        )
        elements.append(Paragraph(self.company_name, company_style))

        # Report title
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2d3748'),
            alignment=TA_CENTER,
            spaceAfter=4
        )
        elements.append(Paragraph(title, title_style))

        # Report type and date
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.gray,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        date_str = datetime.now().strftime("%B %d, %Y")
        elements.append(Paragraph(f"{report_type} | Generated: {date_str}", subtitle_style))

        # Horizontal line
        elements.append(HRFlowable(
            width="100%",
            thickness=1,
            color=colors.HexColor('#3182ce'),
            spaceAfter=20
        ))

        return elements

    def _create_pdf_table(
        self,
        data: List[Dict[str, Any]],
        columns: List[str]
    ) -> List:
        """Create PDF table from data"""
        elements = []
        styles = getSampleStyleSheet()

        if not data:
            return elements

        # Prepare table data
        table_data = []

        # Header row
        header_row = [self._format_header_name(col) for col in columns]
        table_data.append(header_row)

        # Data rows (limit to prevent very large documents)
        max_rows = min(len(data), 1000)
        for row in data[:max_rows]:
            row_data = []
            for col in columns:
                value = row.get(col, "")
                # Format numbers
                if isinstance(value, (int, float, Decimal)):
                    value = self._format_number(value)
                row_data.append(str(value) if value is not None else "")
            table_data.append(row_data)

        # Create table
        col_count = len(columns)
        col_width = (7.5 * inch) / col_count if col_count > 0 else 1.5 * inch

        table = Table(table_data, colWidths=[col_width] * col_count)

        # Apply styling
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3182ce')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),

            # Alternating row colors
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),

            # Row striping
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f7fafc'))
              for i in range(2, len(table_data), 2)],
        ]))

        elements.append(table)

        # Add note if data was truncated
        if len(data) > max_rows:
            note_style = ParagraphStyle(
                'Note',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=TA_CENTER,
                spaceBefore=10
            )
            elements.append(Paragraph(
                f"Showing {max_rows} of {len(data)} records",
                note_style
            ))

        return elements

    def _create_pdf_summary(self, data: List[Dict[str, Any]]) -> List:
        """Create PDF summary section"""
        elements = []
        styles = getSampleStyleSheet()

        elements.append(Spacer(1, 30))
        elements.append(HRFlowable(
            width="100%",
            thickness=0.5,
            color=colors.HexColor('#e2e8f0'),
            spaceAfter=15
        ))

        # Summary title
        summary_style = ParagraphStyle(
            'SummaryTitle',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#2d3748'),
            spaceAfter=10
        )
        elements.append(Paragraph("Summary Statistics", summary_style))

        # Calculate summary statistics
        numeric_columns = []
        if data:
            sample = data[0]
            for key, value in sample.items():
                if isinstance(value, (int, float, Decimal)):
                    numeric_columns.append(key)

        summary_data = [["Metric", "Value"]]
        summary_data.append(["Total Records", str(len(data))])

        for col in numeric_columns[:5]:  # Limit to 5 numeric columns
            values = [row.get(col, 0) for row in data if isinstance(row.get(col), (int, float))]
            if values:
                total = sum(values)
                avg = total / len(values) if values else 0
                max_val = max(values)
                min_val = min(values)

                formatted_name = self._format_header_name(col)
                summary_data.append([f"{formatted_name} - Total", self._format_number(total)])
                summary_data.append([f"{formatted_name} - Average", self._format_number(avg)])
                summary_data.append([f"{formatted_name} - Maximum", self._format_number(max_val)])
                summary_data.append([f"{formatted_name} - Minimum", self._format_number(min_val)])

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#edf2f7')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ]))

        elements.append(summary_table)

        return elements

    def _add_page_footer(self, canvas, doc):
        """Add footer with page numbers"""
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.gray)

        # Page number
        page_num = canvas.getPageNumber()
        text = f"Page {page_num}"
        canvas.drawRightString(
            doc.pagesize[0] - 0.75*inch,
            0.5*inch,
            text
        )

        # Generation timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
        canvas.drawString(
            0.75*inch,
            0.5*inch,
            f"Generated: {timestamp}"
        )

        # Confidential watermark
        canvas.setFillColor(colors.HexColor('#e2e8f0'))
        canvas.setFont('Helvetica', 7)
        canvas.drawCentredString(
            doc.pagesize[0] / 2,
            0.3*inch,
            f"{self.company_name} - Confidential"
        )

        canvas.restoreState()

    def _format_header_name(self, name: str) -> str:
        """Format column header name for display"""
        # Convert snake_case to Title Case
        words = name.replace('_', ' ').split()
        return ' '.join(word.capitalize() for word in words)

    def _format_number(self, value: Union[int, float, Decimal]) -> str:
        """Format number for display"""
        if isinstance(value, Decimal):
            value = float(value)
        return f"{value:,.2f}"

    async def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        title: str,
        sheet_name: str = "Report",
        columns: Optional[List[str]] = None,
        include_charts: bool = True,
        include_summary: bool = True
    ) -> bytes:
        """
        Export data to an Excel workbook

        Args:
            data: List of dictionaries containing report data
            title: Report title
            sheet_name: Name for the worksheet
            columns: Optional list of column names to include
            include_charts: Whether to include charts
            include_summary: Whether to include summary sheet

        Returns:
            Excel workbook as bytes
        """
        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()

        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        # Create data sheet
        data_sheet = workbook.create_sheet(sheet_name)

        if columns is None and data:
            columns = list(data[0].keys())

        # Write headers
        if columns:
            for col_idx, col_name in enumerate(columns, start=1):
                cell = data_sheet.cell(row=1, column=col_idx)
                cell.value = self._format_header_name(col_name)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="3182CE", end_color="3182CE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_idx, row in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = data_sheet.cell(row=row_idx, column=col_idx)
                value = row.get(col_name, "")

                # Format numbers
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                elif isinstance(value, Decimal):
                    cell.value = float(value)
                    cell.number_format = '#,##0.00'

                cell.value = value if value is not None else ""

                # Apply alternating row styling
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(self._format_header_name(col_name))

            for row_idx in range(2, min(len(data) + 2, 100)):  # Check first 100 rows
                cell_value = str(data_sheet.cell(row=row_idx, column=col_idx).value or "")
                max_length = max(max_length, len(cell_value))

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            data_sheet.column_dimensions[column_letter].width = adjusted_width

        # Add charts if requested
        if include_charts and data and columns:
            self._add_excel_charts(workbook, data_sheet, data, columns)

        # Add summary sheet
        if include_summary and data:
            self._add_excel_summary_sheet(workbook, data)

        # Add title sheet
        self._add_title_sheet(workbook, title, len(data))

        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _add_title_sheet(self, workbook, title: str, row_count: int):
        """Add a title/cover sheet"""
        title_sheet = workbook.create_sheet("Cover", 0)

        # Title
        title_sheet.merge_cells('A1:F1')
        title_cell = title_sheet['A1']
        title_cell.value = self.company_name
        title_cell.font = Font(bold=True, size=24, color="1a365d")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_sheet.row_dimensions[1].height = 40

        # Report title
        title_sheet.merge_cells('A3:F3')
        subtitle_cell = title_sheet['A3']
        subtitle_cell.value = title
        subtitle_cell.font = Font(bold=True, size=18, color="2d3748")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Report info
        title_sheet['A5'] = "Report Generated:"
        title_sheet['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        title_sheet['A6'] = "Total Records:"
        title_sheet['B6'] = row_count

        # Styling for info
        for row in [5, 6]:
            title_sheet[f'A{row}'].font = Font(bold=True)
            title_sheet[f'B{row}'].font = Font(size=12)

        # Column widths
        title_sheet.column_dimensions['A'].width = 20
        title_sheet.column_dimensions['B'].width = 30

    def _add_excel_charts(
        self,
        workbook,
        data_sheet,
        data: List[Dict[str, Any]],
        columns: List[str]
    ):
        """Add charts to the Excel sheet"""
        # Find numeric columns for charts
        numeric_cols = []
        for col_idx, col_name in enumerate(columns, start=1):
            sample_values = [row.get(col_name) for row in data[:10] if row.get(col_name) is not None]
            if all(isinstance(v, (int, float, Decimal)) for v in sample_values):
                numeric_cols.append((col_idx, col_name))

        if not numeric_cols:
            return

        # Create a chart sheet
        chart_sheet = workbook.create_sheet("Charts")

        # Bar chart for first numeric column
        if numeric_cols:
            col_idx, col_name = numeric_cols[0]

            chart = BarChart()
            chart.type = "col"
            chart.title = f"{self._format_header_name(col_name)} by Category"
            chart.y_axis.title = self._format_header_name(col_name)
            chart.x_axis.title = "Category"

            # Use first 20 rows for chart
            data_ref = Reference(data_sheet, min_col=col_idx, min_row=1, max_row=min(21, len(data) + 1))
            cats = Reference(data_sheet, min_col=1, min_row=2, max_row=min(21, len(data) + 1))

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4

            chart_sheet.add_chart(chart, "A1")

        # Add summary chart if we have multiple numeric columns
        if len(numeric_cols) >= 2:
            chart2 = BarChart()
            chart2.type = "bar"
            chart2.title = "Comparison of Numeric Values"

            data_ref = Reference(data_sheet, min_col=2, min_row=1, max_col=min(len(numeric_cols) + 1, 5), max_row=min(11, len(data) + 1))
            cats = Reference(data_sheet, min_col=1, min_row=2, max_row=min(11, len(data) + 1))

            chart2.add_data(data_ref, titles_from_data=True)
            chart2.set_categories(cats)

            chart_sheet.add_chart(chart2, "K1")

    def _add_excel_summary_sheet(
        self,
        workbook,
        data: List[Dict[str, Any]]
    ):
        """Add a summary statistics sheet"""
        summary_sheet = workbook.create_sheet("Summary")

        # Title
        summary_sheet['A1'] = "Summary Statistics"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        summary_sheet.merge_cells('A1:C1')

        # Find numeric columns
        numeric_cols = []
        if data:
            for key, value in data[0].items():
                if isinstance(value, (int, float, Decimal)):
                    numeric_cols.append(key)

        row = 3
        summary_sheet[f'A{row}'] = "Metric"
        summary_sheet[f'B{row}'] = "Value"
        summary_sheet[f'A{row}'].font = Font(bold=True)
        summary_sheet[f'B{row}'].font = Font(bold=True)

        row += 1
        summary_sheet[f'A{row}'] = "Total Records"
        summary_sheet[f'B{row}'] = len(data)

        for col in numeric_cols[:5]:
            values = [row.get(col, 0) for row in data if isinstance(row.get(col), (int, float, Decimal))]
            if values:
                row += 1
                summary_sheet[f'A{row}'] = f"{self._format_header_name(col)} - Total"
                summary_sheet[f'B{row}'] = sum(values)
                summary_sheet[f'B{row}'].number_format = '#,##0.00'

                row += 1
                summary_sheet[f'A{row}'] = f"{self._format_header_name(col)} - Average"
                summary_sheet[f'B{row}'] = sum(values) / len(values)
                summary_sheet[f'B{row}'].number_format = '#,##0.00'

                row += 1
                summary_sheet[f'A{row}'] = f"{self._format_header_name(col)} - Max"
                summary_sheet[f'B{row}'] = max(values)
                summary_sheet[f'B{row}'].number_format = '#,##0.00'

                row += 1
                summary_sheet[f'A{row}'] = f"{self._format_header_name(col)} - Min"
                summary_sheet[f'B{row}'] = min(values)
                summary_sheet[f'B{row}'].number_format = '#,##0.00'

        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 20

    async def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        delimiter: str = ','
    ) -> str:
        """
        Export data to CSV format

        Args:
            data: List of dictionaries containing report data
            columns: Optional list of column names to include
            delimiter: CSV delimiter character

        Returns:
            CSV string
        """
        if columns is None and data:
            columns = list(data[0].keys())

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, delimiter=delimiter)

        # Write header
        writer.writeheader()

        # Write data rows
        for row in data:
            filtered_row = {k: row.get(k, '') for k in columns}
            writer.writerow(filtered_row)

        return output.getvalue()

    async def export_financial_statement(
        self,
        statement_type: str,
        data: Dict[str, Any],
        date_range: str,
        format: str = "pdf"
    ) -> Union[bytes, str]:
        """
        Export a formal financial statement (Income Statement, Balance Sheet, Cash Flow)

        Args:
            statement_type: Type of financial statement
            data: Dictionary containing statement data
            date_range: Date range string
            format: Export format (pdf, excel, csv)

        Returns:
            Exported document
        """
        # Convert data to list format for export
        export_data = []

        if statement_type.lower() == "income statement":
            # Revenue section
            for item in data.get('revenue', []):
                export_data.append({
                    'Category': 'Revenue',
                    'Item': item.get('name', ''),
                    'Amount': item.get('amount', 0),
                    'YTD': item.get('ytd', 0)
                })
            # Expenses section
            for item in data.get('expenses', []):
                export_data.append({
                    'Category': 'Expenses',
                    'Item': item.get('name', ''),
                    'Amount': item.get('amount', 0),
                    'YTD': item.get('ytd', 0)
                })
            # Net Income
            export_data.append({
                'Category': 'Summary',
                'Item': 'Net Income',
                'Amount': data.get('net_income', 0),
                'YTD': data.get('net_income_ytd', 0)
            })

        elif statement_type.lower() == "balance sheet":
            # Assets
            for item in data.get('assets', []):
                export_data.append({
                    'Category': 'Assets',
                    'Item': item.get('name', ''),
                    'Current': item.get('current', 0),
                    'Non-Current': item.get('non_current', 0),
                    'Total': item.get('total', 0)
                })
            # Liabilities
            for item in data.get('liabilities', []):
                export_data.append({
                    'Category': 'Liabilities',
                    'Item': item.get('name', ''),
                    'Current': item.get('current', 0),
                    'Non-Current': item.get('non_current', 0),
                    'Total': item.get('total', 0)
                })
            # Equity
            for item in data.get('equity', []):
                export_data.append({
                    'Category': 'Equity',
                    'Item': item.get('name', ''),
                    'Amount': item.get('amount', 0)
                })

        elif statement_type.lower() == "cash flow":
            # Operating activities
            for item in data.get('operating', []):
                export_data.append({
                    'Category': 'Operating Activities',
                    'Item': item.get('name', ''),
                    'Amount': item.get('amount', 0)
                })
            # Investing activities
            for item in data.get('investing', []):
                export_data.append({
                    'Category': 'Investing Activities',
                    'Item': item.get('name', ''),
                    'Amount': item.get('amount', 0)
                })
            # Financing activities
            for item in data.get('financing', []):
                export_data.append({
                    'Category': 'Financing Activities',
                    'Item': item.get('name', ''),
                    'Amount': item.get('amount', 0)
                })

        # Export based on format
        if format.lower() == 'pdf':
            return await self.export_to_pdf(
                export_data,
                title=f"{statement_type.replace('_', ' ').title()}",
                report_type=f"Period: {date_range}",
                include_summary=False
            )
        elif format.lower() == 'excel':
            return await self.export_to_excel(
                export_data,
                title=f"{statement_type.replace('_', ' ').title()}",
                sheet_name=statement_type[:31].replace('_', ' ')  # Excel sheet name limit
            )
        else:
            return await self.export_to_csv(export_data)


# Factory function for creating export service
def create_export_service(company_name: str = "FinAcc", logo_path: Optional[str] = None) -> ExportService:
    """Create an ExportService instance"""
    return ExportService(company_name=company_name, logo_path=logo_path)